#!/usr/bin/env python3
"""
wheel_screener.py -- Income options screener (cash-secured puts + covered calls).

Data source: TRADIER (reliable on servers). Set an access token via env var
TRADIER_TOKEN (sandbox token is free). Yahoo is used only as a light fallback
for VIX and earnings dates.

Run locally:  set TRADIER_TOKEN=your_token   then   py wheel_screener.py
On Streamlit: put the token in the app's Secrets (see DEPLOY.md).
Not financial advice. Screens candidates; you decide.
"""

import os
import json
import math
import time as _time
import datetime as dt

import pandas as pd
from scipy.stats import norm

# ============================== CONFIG ==============================
PUT_TICKERS = ["SPY", "QQQ", "DIA", "MSFT", "GOOG", "NVDA", "AVGO", "SMH", "MRVL", "MU", "IGV",
               "V", "AAPL", "LLY", "MA", "UBER", "NOW", "EWY", "AMZN", "META", "INTC",
               "TSM", "ORCL", "SPCX", "SKHY", "CRWD"]

HOLDINGS = {
    "DKNG": 18.4505,
    "KHC": 32.9834,
    "CMCSA": 35.12,
    "PS": 25.25,
    "NFLX": 69.30,
    "MRVL": 232.03,
    "AVGO": 381.07,
    "EIX": 64.43,
    "GOOG": 359.91,
    "NVDA": 203.925,
    "SMH": 573.6451,
}

# Open positions you've SOLD to open (cash-secured puts, covered calls, credit
# spreads) -- tracked at the bottom of the app: entry price, current cost to
# close, day $/%, and unrealized G/L. All assumed short (you collected a
# credit); G/L = entry_credit - current cost to buy back. "type" is one of
# "put", "call", "put_spread", "call_spread". Single-leg entries need "strike";
# spreads need "short_strike" and "long_strike" instead. See positions.py.
OPEN_POSITIONS = [
    {"ticker": "EIX", "type": "call", "strike": 75, "expiration": "2026-09-18",
     "contracts": 8, "entry_credit": 1.55, "entry_date": "2026-08-07"},
    {"ticker": "KHC", "type": "call", "strike": 27.5, "expiration": "2027-03-19",
     "contracts": 30, "entry_credit": 2.50, "entry_date": "2026-07-29"},
    {"ticker": "MSFT", "type": "call", "strike": 500, "expiration": "2026-08-28",
     "contracts": 3, "entry_credit": 9.75, "entry_date": "2026-08-03"},
    {"ticker": "UBER", "type": "put_spread", "short_strike": 67.5, "long_strike": 65,
     "expiration": "2026-09-18", "contracts": 107, "entry_credit": 0.22, "entry_date": "2026-08-14"},
    {"ticker": "V", "type": "put", "strike": 350, "expiration": "2026-09-18",
     "contracts": 1, "entry_credit": 3.80, "entry_date": "2026-08-17"},
    {"ticker": "MA", "type": "put", "strike": 545, "expiration": "2026-09-18",
     "contracts": 1, "entry_credit": 6.4999, "entry_date": "2026-08-17"},
    # Iron condor -- tracked as separate put_spread/call_spread entries since
    # OPEN_POSITIONS has no combined iron_condor type. Put side opened
    # 08/19/2026; call side (below) completed it on 08/24/2026.
    {"ticker": "GOOG", "type": "put_spread", "short_strike": 305, "long_strike": 290,
     "expiration": "2026-09-18", "contracts": 14, "entry_credit": 0.739971, "entry_date": "2026-08-19"},
    {"ticker": "SMH", "type": "put_spread", "short_strike": 505, "long_strike": 480,
     "expiration": "2026-09-18", "contracts": 12, "entry_credit": 2.879875, "entry_date": "2026-08-20"},
    {"ticker": "SPCX", "type": "call_spread", "short_strike": 150, "long_strike": 155,
     "expiration": "2026-09-18", "contracts": 30, "entry_credit": 1.049913, "entry_date": "2026-08-21"},
    {"ticker": "SPCX", "type": "call_spread", "short_strike": 160, "long_strike": 170,
     "expiration": "2026-10-16", "contracts": 15, "entry_credit": 1.699887, "entry_date": "2026-08-21"},
    {"ticker": "TSM", "type": "put", "strike": 380, "expiration": "2026-09-18",
     "contracts": 2, "entry_credit": 4.8999, "entry_date": "2026-08-24"},
    # LLY 1150P -- two tax lots tracked together as one averaged position:
    # (18.9996 + 20.9996) / 2 contracts = 19.9996/share. entry_date kept as
    # the earlier lot's (08/26), when the position was first opened.
    {"ticker": "LLY", "type": "put", "strike": 1150, "expiration": "2026-09-18",
     "contracts": 2, "entry_credit": 19.9996, "entry_date": "2026-08-26"},
    {"ticker": "SKHY", "type": "put_spread", "short_strike": 145, "long_strike": 138,
     "expiration": "2026-09-04", "contracts": 39, "entry_credit": 0.699964, "entry_date": "2026-08-27"},
    {"ticker": "ORCL", "type": "call_spread", "short_strike": 167.5, "long_strike": 175,
     "expiration": "2026-09-04", "contracts": 35, "entry_credit": 0.34998, "entry_date": "2026-08-27"},
    {"ticker": "MRVL", "type": "call", "strike": 250, "expiration": "2026-09-18",
     "contracts": 2, "entry_credit": 4.35, "entry_date": "2026-08-28"},
    {"ticker": "NOW", "type": "call_spread", "short_strike": 170, "long_strike": 185,
     "expiration": "2026-10-16", "contracts": 20, "entry_credit": 2.30, "entry_date": "2026-08-31"},
    {"ticker": "INTC", "type": "call_spread", "short_strike": 105, "long_strike": 110,
     "expiration": "2026-10-16", "contracts": 56, "entry_credit": 0.55, "entry_date": "2026-09-01"},
    {"ticker": "META", "type": "call_spread", "short_strike": 625, "long_strike": 655,
     "expiration": "2026-09-18", "contracts": 9, "entry_credit": 2.06, "entry_date": "2026-09-01"},
    {"ticker": "AMZN", "type": "put", "strike": 230, "expiration": "2026-10-16",
     "contracts": 2, "entry_credit": 2.40, "entry_date": "2026-09-01"},
    {"ticker": "GOOG", "type": "put", "strike": 300, "expiration": "2026-11-20",
     "contracts": 2, "entry_credit": 6.75, "entry_date": "2026-09-01"},
    {"ticker": "AAPL", "type": "call_spread", "short_strike": 360, "long_strike": 380,
     "expiration": "2026-10-16", "contracts": 14, "entry_credit": 1.74, "entry_date": "2026-09-03"},
    {"ticker": "AVGO", "type": "put", "strike": 320, "expiration": "2026-10-16",
     "contracts": 15, "entry_credit": 4.2706, "entry_date": "2026-09-04"},
]

# Closed positions, same shape as OPEN_POSITIONS plus "exit_cost" (what you paid
# to buy it back -- 0 if it expired worthless / hit max profit) and "exit_date".
# Shown at the bottom of the app for a rolling window (default 30 days) after
# exit_date -- pure arithmetic against the recorded exit price, no live quotes
# needed since the trade is already settled. See positions.py.
CLOSED_POSITIONS = [
    {"ticker": "MU", "type": "call_spread", "short_strike": 1060, "long_strike": 1110,
     "expiration": "2026-08-21", "contracts": 6, "entry_credit": 3.639867, "entry_date": "2026-08-14",
     "exit_cost": 0.24, "exit_date": "2026-08-19"},
    {"ticker": "INTC", "type": "call_spread", "short_strike": 135, "long_strike": 140,
     "expiration": "2026-09-18", "contracts": 53, "entry_credit": 0.34997, "entry_date": "2026-08-17",
     "exit_cost": 0.08, "exit_date": "2026-08-19"},
    {"ticker": "MRVL", "type": "put_spread", "short_strike": 185, "long_strike": 175,
     "expiration": "2026-08-21", "contracts": 28, "entry_credit": 1.270311, "entry_date": "2026-08-11",
     "exit_cost": 0.040004, "exit_date": "2026-08-17"},
    {"ticker": "AMZN", "type": "call_spread", "short_strike": 310, "long_strike": 325,
     "expiration": "2026-09-18", "contracts": 15, "entry_credit": 1.409947, "entry_date": "2026-08-07",
     "exit_cost": 0.19, "exit_date": "2026-08-17"},
    {"ticker": "MU", "type": "put", "strike": 650, "expiration": "2026-08-14",
     "contracts": 1, "entry_credit": 10.3498, "entry_date": "2026-07-30",
     "exit_cost": 0, "exit_date": "2026-08-14"},   # expired worthless, full premium kept
    {"ticker": "AVGO", "type": "call", "strike": 415, "expiration": "2026-08-21",
     "contracts": 1, "entry_credit": 12.0498, "entry_date": "2026-07-21",
     "exit_cost": 2.01, "exit_date": "2026-08-14"},
    {"ticker": "DKNG", "type": "call", "strike": 25, "expiration": "2026-08-21",
     "contracts": 11, "entry_credit": 1.77, "entry_date": "2026-07-17",
     "exit_cost": 0.3795, "exit_date": "2026-08-06"},
    {"ticker": "KHC", "type": "call", "strike": 25, "expiration": "2026-07-31",
     "contracts": 30, "entry_credit": 0.445, "entry_date": "2026-06-17",
     "exit_cost": 2.65, "exit_date": "2026-07-29"},
    {"ticker": "EWY", "type": "put", "strike": 135, "expiration": "2026-08-21",
     "contracts": 3, "entry_credit": 3.4999, "entry_date": "2026-07-31",
     "exit_cost": 0.40, "exit_date": "2026-08-10"},
    # MRVL 250C -- two separate tax lots (opened/closed at different times), kept
    # as two entries rather than combined so each reconciles individually.
    {"ticker": "MRVL", "type": "call", "strike": 250, "expiration": "2026-08-21",
     "contracts": 2, "entry_credit": 7.7999, "entry_date": "2026-08-04",
     "exit_cost": 1.30, "exit_date": "2026-08-11"},
    {"ticker": "MRVL", "type": "call", "strike": 250, "expiration": "2026-08-21",
     "contracts": 2, "entry_credit": 10.2998, "entry_date": "2026-07-21",
     "exit_cost": 1.65, "exit_date": "2026-07-28"},
    {"ticker": "DKNG", "type": "call", "strike": 27.5, "expiration": "2026-09-18",
     "contracts": 11, "entry_credit": 0.668782, "entry_date": "2026-08-21",
     "exit_cost": 0.138, "exit_date": "2026-09-01"},
    {"ticker": "NVDA", "type": "call_spread", "short_strike": 250, "long_strike": 260,
     "expiration": "2026-09-18", "contracts": 27, "entry_credit": 0.714041, "entry_date": "2026-08-27",
     "exit_cost": 0.11, "exit_date": "2026-09-01"},
    {"ticker": "GOOG", "type": "call_spread", "short_strike": 385, "long_strike": 405,
     "expiration": "2026-09-18", "contracts": 14, "entry_credit": 0.839971, "entry_date": "2026-08-24",
     "exit_cost": 0.054286, "exit_date": "2026-09-04"},
    {"ticker": "EWY", "type": "put_spread", "short_strike": 145, "long_strike": 140,
     "expiration": "2026-09-18", "contracts": 45, "entry_credit": 1.129902, "entry_date": "2026-08-10",
     "exit_cost": 0.042333, "exit_date": "2026-09-04"},
]

# 70% POP anchor (Options Alpha): POP = 1 - |delta|, so ~70% POP ~= 0.30 delta.
POP_MIN           = 0.70     # at least 70% POP (delta <= ~0.30); higher POP is fine too
POP_MAX           = 1.0      # no upper cap
DTE_MIN           = 7      # include short weeklies
DTE_MAX           = 60     # no long-dated contracts
YIELD_HURDLE_BASE = 0.25     # (informational; the active yield rule is the two lines below)
MIN_ANN_YIELD     = 0.15     # flat floor: contracts must pay >= this annualized (when tiered rule off)
MIN_ANN_YIELD_INDEX = 0.10   # broad indexes (SPY/QQQ/DIA) are lower risk -> lower yield floor is OK
MIN_PERIOD_YIELD  = 0.01     # require at least 1% period (per-contract) yield
MIN_OPEN_INTEREST = 1000     # minimum open interest for a contract/leg to appear (0 to disable)
CASH_TARGET = 40000          # capital target; screener shows # of contracts to reach at least this

# --- Cash-secured-put-only filters (do NOT apply to covered calls or spreads) ---
PUT_MIN_PREMIUM      = 0.0    # absolute $/share premium floor. OFF (using % of strike below instead).
PUT_MIN_PREMIUM_PCT  = 0.015  # premium must be >= this fraction of the strike (1.5% of strike). 0 to disable.
PUT_MIN_YIELD_OVER_IV = 0.0   # annualized yield >= this fraction of IV. OFF this round.
PUT_MIN_OTM_OVER_IV  = 0.15   # puts: OTM distance must be >= this fraction of IV. 0 to disable.
CALL_MIN_OTM_OVER_IV = 0.15   # covered calls: OTM distance must be >= this fraction of IV. 0 to disable.
# Diversity: collapse each ticker's strike/expiry ladder to the single best-SCORE
# contract PER EXPIRATION. Cuts a dominant ticker's row count without a hard cap.
PUT_BEST_PER_EXPIRATION  = True
CALL_BEST_PER_EXPIRATION = True
# Composite ranking score (puts AND covered calls): Score = (AnnYield / IV^c) * POP^a * (365/DTE)^b.
# Dividing by IV strips out the fact that yield is naturally richer on higher-IV (riskier) names,
# so the score no longer rewards volatility. POP and shorter DTE break ties.
SCORE_POP_EXP = 1.0   # a: how much higher POP is rewarded (0 = ignore POP)
SCORE_DTE_EXP = 0.5   # b: how much shorter DTE is rewarded (0 = ignore length)
SCORE_IV_EXP  = 1.0   # c: IV penalty. 0 = ignore IV (old behavior), 1 = neutralize IV, >1 = actively favor calmer names
USE_TIERED_YIELD  = False    # ON: use the tiered OTM->yield rule below instead of the flat floor
TIERED_YIELD = [(0.15, 0.10), (0.10, 0.15), (0.05, 0.25)]  # (min OTM, required ann. yield), high OTM first
DTE_SHORT_CUTOFF    = 21     # <=21 days = "3 weeks and under"
YIELD_OVER_IV_SHORT = 1.0    #   short-dated (<=21 DTE): annualized yield must be > 100% of IV
YIELD_OVER_IV_LONG  = 0.7    # 22+ DTE: annualized yield must be > 70% of IV
USE_YIELD_OVER_IV   = False  # OFF: don't require yield to beat IV (set True to re-enable)
REQUIRE_STRIKE_ABOVE_COST = False  # OFF: covered-call strike need NOT be above your cost basis
INDEX_TICKERS       = {"SPY", "QQQ", "DIA"}   # these get the 5% OTM floor; all others 10%
OTM_MIN_INDEX       = 0.05   # min % OTM for index ETFs (applies to ALL strategies)
OTM_MIN_OTHER       = 0.10   # min % OTM for every other ticker (applies to ALL strategies)
OTM_MAX             = 1.0    # max % OTM for single-leg (1.0 = effectively off)
NO_EARNINGS_TICKERS = {"SPY", "QQQ", "DIA", "SMH", "IGV", "EWY"}   # ETFs: no earnings to span
EXCLUDE_IF_EARNINGS_UNKNOWN = False  # show stocks even if earnings date unconfirmed (use EARNINGS_DATES to be safe)
ALLOW_EARNINGS_IN_WINDOW = False  # True: DO show contracts whose window spans an earnings date. False: exclude them.
# Manual earnings dates (YYYY-MM-DD). These OVERRIDE Yahoo and are the most reliable filter.
# Fill in your tickers' next earnings date; update roughly once a quarter.
EARNINGS_DATES = {
    "UBER": "2026-08-05",
    # add your other stocks' next earnings the same way (YYYY-MM-DD):
    # "NVDA": "2026-08-27",
    # "MSFT": "2026-07-29",
    # "LLY":  "2026-08-06",
}

# Related tickers whose OWN earnings can be a real catalyst for this one, even
# with nothing on this ticker's own calendar (e.g. a chip equipment maker's
# earnings moving a memory maker via demand read-through, or a direct
# competitor's results). Used only by the long strangle/straddle screener
# (spreads.py) -- see _catalyst_dates -- everything else here is unaffected.
# Fully automatic (checked live via get_earnings_date, same as your own
# earnings), no manual date entry needed; edit the peer LIST itself if a
# relationship changes.
#
# Each entry is either evidence-backed (a real, sourced example of one
# actually moving the other around earnings -- same direction or, for direct
# competitors, opposite/divergent moves that still create real volatility)
# or a genuine head-to-head competitor even without a specific correlation
# example on record. Entries considered and dropped for lacking either:
# V/MA's cross-list to AXP/PYPL ("earnings announcements and stock movements
# don't necessarily correlate strongly with each other" -- explicit evidence
# against it), AAPL's old TSM/AVGO guess (no evidence either way), LLY's old
# PFE/MRK (not direct GLP-1 competitors the way NVO is), EIX's old SO/DUK/SRE
# (regulated utilities don't compete for customers, and no correlation
# evidence was found -- only PCG had real supporting evidence, via shared
# California wildfire liability dynamics).
PEER_TICKERS = {
    "MSFT": ["GOOG", "AMZN", "META", "AAPL"],   # "move together in virtually all cases" -- mega-cap
    "GOOG": ["MSFT", "AMZN", "META", "AAPL"],   # tech correlation, sourced; all four (+AAPL) also
    "AMZN": ["MSFT", "GOOG", "META", "AAPL"],   # trace back to the same Samsung/SK Hynix/Micron
    "META": ["MSFT", "GOOG", "AMZN", "AAPL"],   # HBM memory bottleneck -- ties this cluster to MU/NVDA too
    "AAPL": ["MSFT", "GOOG", "AMZN", "META"],
    "NVDA": ["AMD", "AVGO", "TSM"],   # TSM: "NVIDIA is the single most Taiwan Semiconductor-
                                      # dependent name" (supply read-through, sourced); AMD: direct
                                      # AI-chip competitor, but only a moderate/mixed 0.53 stock
                                      # correlation, sometimes opposite-direction on earnings --
                                      # kept for the direct competition, not the correlation; AVGO:
                                      # direct competitor. (MU removed -- not a competitor, different
                                      # products; despite a real commercial link -- Micron supplies
                                      # HBM memory for NVDA's GPUs -- the stock evidence actually
                                      # shows divergence, not correlation: "Since earnings, Micron
                                      # surged 128% while NVDA fell 10%")
    "AVGO": ["NVDA", "MRVL", "QCOM"],   # AVGO/MRVL: "intense rivals" in AI networking chips, sourced
    "MRVL": ["AVGO", "NVDA", "AMD", "TSM"],
    "MU":   ["AMAT", "ASML", "LRCX", "KLAC",   # equipment makers: sourced, all 4 jumped 4-6%+
             "SKHY", "SSNLF", "SNDK"],         # specifically on MU's earnings. SK Hynix/Samsung/
                                               # SanDisk: strongly sourced -- "SK Hynix fell 6%...
                                               # Micron also declined 6%" (near-identical synchronized
                                               # moves) with SanDisk down 9% in the same DRAM-pricing
                                               # selloff; explicitly because the three renegotiate
                                               # DRAM/NAND pricing off shared supply-demand dynamics.
                                               # (TSM removed -- confirmed NOT a competitor, "they do
                                               # not compete because they are making different types
                                               # of chips" -- logic vs. memory -- and no correlation
                                               # evidence either; memory is far more cyclical/commodity-
                                               # priced than logic, and the two stocks have diverged
                                               # sharply, TSM +~400% vs MU +~700% since 2023)
    "INTC": ["AMD", "QCOM", "TSM", "SSNLF", "NVDA"],   # AMD: direct x86 CPU rival, the classic
                                                        # INTC/AMD rivalry; QCOM: Snapdragon X now
                                                        # directly competing for Windows PC/laptop
                                                        # CPU share ("Qualcomm Snapdragon Takes Aim
                                                        # At PC Rivals Intel..."); TSM/SSNLF: Intel
                                                        # Foundry vs. TSMC/Samsung Foundry, grouped
                                                        # industry-wide as the "big three" advanced-
                                                        # node foundries; NVDA: still a direct data-
                                                        # center/AI-compute rival despite NVIDIA's
                                                        # $5B equity stake + joint roadmap (Sept 2025)
                                                        # -- coverage calls it "both partner and
                                                        # competitor." (ARM considered and excluded --
                                                        # its March 2026 Arm AGI CPU is its first own-
                                                        # brand chip and a genuine data-center
                                                        # competitor, but too new/narrow a product to
                                                        # be a real driver of INTC moves yet)
    "V":    ["MA"],   # direct card-network duopoly competitors
    "MA":   ["V"],
    "LLY":  ["NVO"],   # sourced: LLY +7% / NVO -6% same week off each other's GLP-1 results --
                       # opposite-direction, but real, evidenced cross-reactivity
    "UBER": ["LYFT", "DASH"],   # direct rideshare/delivery competitors (LYFT's stock correlation
                                # is actually low, ~0.49 -- kept anyway as a direct competitor)
    "NOW":  ["CRM", "WDAY"],   # direct enterprise-software competitors
    "DKNG": ["FLUT", "MGM", "CZR"],   # sourced: DKNG -8%, "selling off in sympathy with Flutter's
                                      # 14% post-earnings drop"; MGM/CZR: direct competitors
    "KHC":  ["GIS", "MDLZ"],   # direct packaged-food competitors (Kellanova/K delisted Dec 2025 --
                               # acquired by Mars, removed)
    "CMCSA": ["CHTR", "DIS", "WBD"],   # direct cable/streaming competitors
    "NFLX": ["DIS", "WBD", "PSKY"],   # direct streaming competitors (Paramount now trades as PSKY,
                                      # not PARA, post-Skydance merger)
    "EIX":  ["PCG"],   # same California wildfire liability dynamics, explicitly cross-referenced
                       # in coverage of both -- not a competitor (regulated monopoly), a shared-risk peer
    # PS (Pershing Square Inc.) intentionally has no entry -- it's a holding
    # company driven by its own portfolio's performance, not sector peers.
}
USE_TBILL_SPREAD  = False    # your old "beat T-bill by 5pts" rule (off; set True to re-enable)
MIN_RISK_PREMIUM  = 0.05
IVR_MIN           = 0.50
USE_IVR           = False    # Tradier gives current IV but NOT IV rank (needs IV history)

RISK_FREE         = 0.043
PREMIUM_BASIS     = "bid"
TBILL_LADDER = [(35, 0.0370), (56, 0.0372), (100, 0.0379), (190, 0.0390), (99999, 0.0398)]
OUTPUT_FILE = "qualifying_contracts.xlsx"
# ===================================================================


# ------------------------- Tradier data ---------------------------
def _td_get(path, params):
    import requests
    token = os.environ.get("TRADIER_TOKEN", "")
    base = os.environ.get("TRADIER_BASE", "https://sandbox.tradier.com/v1")
    r = requests.get(base + path, params=params, timeout=20,
                     headers={"Authorization": "Bearer " + token,
                              "Accept": "application/json"})
    r.raise_for_status()
    return r.json()


def _as_list(x):
    if x is None:
        return []
    return x if isinstance(x, list) else [x]


def td_quote(symbol):
    j = _td_get("/markets/quotes", {"symbols": symbol})
    q = (j.get("quotes") or {}).get("quote")
    q = q[0] if isinstance(q, list) else q
    if not q:
        return None
    return q.get("last") or q.get("close") or q.get("prevclose")


def td_expirations(symbol):
    j = _td_get("/markets/options/expirations",
                {"symbol": symbol, "includeAllRoots": "true"})
    return _as_list((j.get("expirations") or {}).get("date"))


def td_chain(symbol, expiration):
    j = _td_get("/markets/options/chains",
                {"symbol": symbol, "expiration": expiration, "greeks": "true"})
    out = []
    for o in _as_list((j.get("options") or {}).get("option")):
        g = o.get("greeks") or {}
        out.append({"type": o.get("option_type"),
                    "strike": float(o.get("strike")),
                    "bid": o.get("bid") or 0,
                    "ask": o.get("ask") or 0,
                    "prevclose": o.get("prevclose") or 0,
                    "delta": g.get("delta"),
                    "iv": g.get("mid_iv") or g.get("smv_vol") or 0,
                    "oi": o.get("open_interest") or 0,
                    "volume": o.get("volume") or 0})
    return out


# ------------------- Yahoo fallback (VIX/earnings only) -----------
def _make_session():
    try:
        from curl_cffi import requests as _cffi
        return _cffi.Session(impersonate="chrome")
    except Exception:
        return None


_SESSION = _make_session()


def _ticker(sym):
    import yfinance as yf
    try:
        return yf.Ticker(sym, session=_SESSION) if _SESSION else yf.Ticker(sym)
    except Exception:
        return yf.Ticker(sym)


def tbill_for(dte):
    for max_dte, rate in TBILL_LADDER:
        if dte <= max_dte:
            return rate
    return TBILL_LADDER[-1][1]


def get_vix():
    try:
        q = td_quote("VIX")
        if q:
            return round(float(q), 2)
    except Exception:
        pass
    try:
        v = _ticker("^VIX")
        px = v.history(period="1d")["Close"].iloc[-1]
        return round(float(px), 2)
    except Exception:
        return None


def vix_regime(v):
    if v is None:
        return "unknown"
    if v < 15:
        return "low - thin premiums, mostly wait"
    if v < 20:
        return "below-average - selective"
    if v < 30:
        return "elevated - prime selling"
    return "high - rich but risky"


def _d1(S, K, T, sigma, r):
    return (math.log(S / K) + (r + 0.5 * sigma ** 2) * T) / (sigma * math.sqrt(T))


def bs_put_delta(S, K, T, sigma, r):
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return float("nan")
    return norm.cdf(_d1(S, K, T, sigma, r)) - 1.0


def bs_call_delta(S, K, T, sigma, r):
    if T <= 0 or sigma <= 0 or S <= 0 or K <= 0:
        return float("nan")
    return norm.cdf(_d1(S, K, T, sigma, r))


def otm_min_for(symbol):
    """Per-ticker minimum OTM: 5% for index ETFs, 10% for everything else."""
    return OTM_MIN_INDEX if symbol in INDEX_TICKERS else OTM_MIN_OTHER


def _load_history():
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "history_premiums.json")
        with open(path) as f:
            return json.load(f)
    except Exception:
        return {"_meta": {}, "tickers": {}}


_HISTORY = _load_history()
# build_history.py switched from storing a $ premium band to an annualized-yield
# fraction band (comparable to AnnYield_%/AnnROR_% directly). A file built before
# that change (or missing this marker) is treated as unavailable rather than
# misread -- a $8.23 read as a 823% yield would be a silent, dangerous error.
_HISTORY_METRIC_OK = _HISTORY.get("_meta", {}).get("metric") == "ann_yield"


def load_volume_leaders():
    """Daily background scan of high-options-volume S&P 500 tickers outside
    PUT_TICKERS (see build_volume_leaders.py). None if the Action hasn't run yet."""
    try:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "volume_leaders.json")
        with open(path) as f:
            return json.load(f)
    except Exception:
        return None


def _nearest(value, buckets):
    return min(buckets, key=lambda b: abs(b - value)) if buckets else value


def _bracket(value, buckets):
    """The two buckets bracketing `value`, for interpolation instead of
    nearest-neighbor snapping -- clamped at the ends (no extrapolation beyond
    the grid). Snapping both a spread's short and long leg to the SAME nearest
    bucket (common when they're close in OTM% -- e.g. a 5%-wide spread against
    a 5-point bucket grid) erases the richness gradient between them, silently
    collapsing the modeled credit toward zero even though the real strikes are
    genuinely different distances from the money."""
    buckets = sorted(buckets)
    if value <= buckets[0]:
        return buckets[0], buckets[0]
    if value >= buckets[-1]:
        return buckets[-1], buckets[-1]
    for lo, hi in zip(buckets, buckets[1:]):
        if lo <= value <= hi:
            return lo, hi
    return buckets[-1], buckets[-1]


def _lerp_band(a, b, frac):
    """Linearly interpolates between two [low, high] bands. Falls back to
    whichever single band is available if only one side has data."""
    if a is None or b is None:
        return a if b is None else b
    return [a[0] + (b[0] - a[0]) * frac, a[1] + (b[1] - a[1]) * frac]


def _interp_band(table, key_lo, key_hi, lo, hi, value):
    """Linearly interpolates a [low, high] band between two bucket keys."""
    frac = 0.0 if hi == lo else (value - lo) / (hi - lo)
    return _lerp_band(table.get(key_lo), table.get(key_hi), frac)


def avg_premium_range(symbol, otm, dte, kind="put", iv=None):
    """Historical typical ANNUALIZED YIELD [low, high] (fractions, e.g. 0.082 =
    8.2%) for this ticker at this OTM%/DTE, interpolated between the bracketing
    buckets in history_premiums.json (not snapped to whichever single bucket is
    nearest -- see _bracket's docstring for why that matters) -- directly
    comparable to AnnYield_%/AnnROR_%. kind is 'put' or 'call'. If `iv` is given
    (the live contract's IV, as a fraction), also tries the band conditioned on
    that vol regime -- "typical yield given today's vol", not averaged across
    every vol regime the ticker saw all year -- falling back to the
    unconditional OTM/DTE band if that finer bucket doesn't exist (either not
    enough historical days at that vol level, or an older file built before
    vol-conditioning existed). None if not available at all."""
    if not _HISTORY_METRIC_OK:
        return None
    tkr = (_HISTORY.get("tickers") or {}).get(symbol)
    if not tkr:
        return None
    # new tables are {"put": {...}, "call": {...}}; old ones were a flat put table
    if "put" in tkr or "call" in tkr:
        table = tkr.get(kind)
    else:
        table = tkr if kind == "put" else None
    if not table:
        return None
    meta = _HISTORY.get("_meta", {})
    otm_pct, dte_val = otm * 100, dte
    ob_lo, ob_hi = _bracket(otm_pct, meta.get("otm_buckets", [5, 10, 15, 20, 25, 30]))
    db_lo, db_hi = _bracket(dte_val, meta.get("dte_buckets", [7, 14, 21, 30, 45, 60, 90]))

    def _lookup(suffix=""):
        # Interpolate DTE first (at each bracketing OTM bucket), then interpolate
        # those two DTE-interpolated results across OTM -- a small bilinear
        # interpolation across the 2D (OTM, DTE) grid.
        def _at_otm(ob):
            return _interp_band(table, f"{int(ob)}|{int(db_lo)}{suffix}",
                                f"{int(ob)}|{int(db_hi)}{suffix}", db_lo, db_hi, dte_val)
        otm_frac = 0.0 if ob_hi == ob_lo else (otm_pct - ob_lo) / (ob_hi - ob_lo)
        return _lerp_band(_at_otm(ob_lo), _at_otm(ob_hi), otm_frac)

    if iv is not None and iv > 0:
        ivb = _nearest(iv * 100, meta.get("iv_buckets", [15, 25, 35, 50, 70, 90, 120]))
        vol_conditioned = _lookup(suffix=f"|{int(ivb)}")
        if vol_conditioned is not None:
            return vol_conditioned
    return _lookup()


def open_position_sides(symbol):
    """Which side(s) -- 'put', 'call' -- already have an open position on this
    ticker, from OPEN_POSITIONS. Used to keep the screener (puts/calls/spreads,
    including Discover) from suggesting more of the same directional exposure
    on a ticker you're already committed to -- e.g. an open INTC put spread
    blocks new INTC puts/put spreads, but INTC call spreads still show.
    Contract Lookup deliberately ignores this -- it's a manual override tool
    for looking up any strike/expiration regardless of screener criteria."""
    sides = set()
    for p in OPEN_POSITIONS:
        if p["ticker"] != symbol:
            continue
        if p["type"] in ("put", "put_spread"):
            sides.add("put")
        elif p["type"] in ("call", "call_spread"):
            sides.add("call")
    return sides


_SECTOR_CACHE = {}
# GICS "Technology" plus "Communication Services" -- GOOG/META land in the
# latter under GICS, but this app's own PEER_TICKERS already treats them as
# part of the same tech cluster as MSFT/AMZN/AAPL, so the Concentration of
# Positions table follows the same judgment rather than the raw GICS split.
TECH_SECTORS = {"Technology", "Communication Services"}

# Manual overrides -- win over both the sector/category lookup below,
# same "manual entry wins" precedent as EARNINGS_DATES. For names where the
# raw GICS/category classification misses the real tech exposure:
#   SPCX: GICS "Industrials" (aerospace) -- but the business is a tech
#     company (reusable rockets, Starlink satellite internet), not a
#     traditional industrial/manufacturing name.
#   EWY:  category "Focused Region" (South Korea), not tech at all by
#     category -- but its two largest holdings, Samsung Electronics and SK
#     Hynix, are both massive AI-memory chipmakers, so its return stream is
#     effectively a memory/AI-chip proxy, not a generic country fund.
#   QQQ:  category "Large Growth" (broad Nasdaq-100 index fund), but more
#     than half its weight sits in mega-cap tech names, so its return
#     stream tracks tech far more than a generic large-cap fund.
#   AMZN: GICS "Consumer Cyclical" -- same quirk as GOOG/META landing in
#     Communication Services, except AMZN doesn't even get that override.
#     This app's own PEER_TICKERS already treats it as part of the same
#     mega-cap tech cluster as MSFT/GOOG/META/AAPL.
SECTOR_OVERRIDES = {"SPCX": "Tech", "EWY": "Tech", "QQQ": "Tech", "AMZN": "Tech"}


def get_sector_bucket(symbol):
    """'Tech' or 'Non-Tech' for this ticker, for the Concentration of
    Positions table. SECTOR_OVERRIDES wins first; otherwise stocks carry a
    GICS "sector" via yfinance, ETFs don't so falls back to "category" (e.g.
    SMH has no sector but category "Technology"). Cached in-process
    (module-level dict) since this barely changes and yfinance's .info call
    is slow -- shared across notify_email.py and the Streamlit app, unlike
    Streamlit's own st.cache_data, which only exists in the app process."""
    if symbol in _SECTOR_CACHE:
        return _SECTOR_CACHE[symbol]
    if symbol in SECTOR_OVERRIDES:
        _SECTOR_CACHE[symbol] = SECTOR_OVERRIDES[symbol]
        return SECTOR_OVERRIDES[symbol]
    bucket = "Non-Tech"
    try:
        info = _ticker(symbol).info
        if (info.get("sector") or info.get("category")) in TECH_SECTORS:
            bucket = "Tech"
    except Exception:
        pass
    _SECTOR_CACHE[symbol] = bucket
    return bucket


def contracts_for_target(cash_per_contract, target=None):
    """How many contracts (whole number) to tie up at least `target` dollars of capital."""
    t = CASH_TARGET if target is None else target
    if not cash_per_contract or cash_per_contract <= 0:
        return float("nan")
    return int(math.ceil(t / cash_per_contract))


def _liq(o):
    """Liquidity snapshot for one option: raw ask (Premium is shown as a bid-ask
    range, so no separate Spread_$ column), open interest, volume."""
    return {"Ask": o.get("ask") or 0,
            "OpenInt": int(o.get("oi") or 0),
            "Volume": int(o.get("volume") or 0)}


def earnings_blocks(symbol, earnings, today, exp_date):
    """True if this expiration window must be excluded because of earnings.
    Known date in window -> block. ETF (no earnings) -> never. Stock w/ unknown date -> block if strict."""
    if ALLOW_EARNINGS_IN_WINDOW:
        return False
    if earnings is not None:
        return today <= earnings <= exp_date
    if symbol in NO_EARNINGS_TICKERS:
        return False
    return EXCLUDE_IF_EARNINGS_UNKNOWN


def tiered_yield_needed(otm):
    """Required annualized yield by OTM bracket (further OTM -> less yield needed)."""
    for min_otm, req in TIERED_YIELD:      # ordered high OTM -> low
        if otm >= min_otm:
            return req
    return TIERED_YIELD[-1][1]              # closer than smallest tier -> strictest


def evaluate_put(row, spot, dte, earnings_in_window, iv_rank=None, delta=None, otm_min=None, is_index=False):
    strike, premium, iv = row["strike"], row["premium"], row["iv"]
    otm     = (spot - strike) / spot
    per_yld = premium / strike if strike else float("nan")
    ann_yld = per_yld * 365.0 / dte if dte else float("nan")
    d = delta if delta is not None else bs_put_delta(spot, strike, dte / 365.0, iv, RISK_FREE)
    absdelta = abs(d)
    delta_pct = 1.0 - absdelta
    tbill    = tbill_for(dte)
    risk_prem = ann_yld - tbill
    needed   = YIELD_HURDLE_BASE - otm
    yiv      = YIELD_OVER_IV_SHORT if dte <= DTE_SHORT_CUTOFF else YIELD_OVER_IV_LONG
    # broad indexes get a lower yield floor (lower risk) and skip the stock-oriented premium floors
    flat_floor = MIN_ANN_YIELD_INDEX if is_index else MIN_ANN_YIELD
    req_yield = tiered_yield_needed(otm) if USE_TIERED_YIELD else flat_floor
    _om = otm_min if otm_min is not None else OTM_MIN_OTHER
    tests = {
        "pop_target":   POP_MIN <= delta_pct <= POP_MAX,
        "min_yield":    ann_yld >= req_yield,
        "min_period_yield": is_index or (per_yld >= MIN_PERIOD_YIELD),
        "dte_window":   DTE_MIN <= dte <= DTE_MAX,
        "otm_range":    _om <= otm <= OTM_MAX,
        "no_earnings":  not earnings_in_window,
    }
    if PUT_MIN_PREMIUM > 0:
        tests["min_premium"] = premium >= PUT_MIN_PREMIUM
    if PUT_MIN_PREMIUM_PCT > 0 and not is_index:
        tests["min_premium_pct"] = per_yld >= PUT_MIN_PREMIUM_PCT
    if PUT_MIN_YIELD_OVER_IV > 0:
        tests["yield_vs_iv"] = bool(iv) and ann_yld >= PUT_MIN_YIELD_OVER_IV * iv
    if PUT_MIN_OTM_OVER_IV > 0:
        tests["otm_vs_iv"] = bool(iv) and otm >= PUT_MIN_OTM_OVER_IV * iv
    if USE_YIELD_OVER_IV:
        tests["yield_over_iv"] = ann_yld > yiv * iv
    if USE_TBILL_SPREAD:
        tests["tbill_spread"] = risk_prem >= MIN_RISK_PREMIUM
    if USE_IVR:
        tests["iv_rank"] = (iv_rank is not None) and (iv_rank >= IVR_MIN)
    reasons = []
    if not tests["pop_target"]:
        reasons.append(f"POP {delta_pct:.0%} outside {POP_MIN:.0%}-{POP_MAX:.0%}")
    if not tests["min_yield"]:
        reasons.append(f"yield {ann_yld:.1%} < {req_yield:.0%} needed")
    if not tests["min_period_yield"]:
        reasons.append(f"period yield {per_yld:.1%} < {MIN_PERIOD_YIELD:.0%}")
    if PUT_MIN_PREMIUM > 0 and not tests.get("min_premium"):
        reasons.append(f"premium ${premium:.2f} < ${PUT_MIN_PREMIUM:.0f}")
    if PUT_MIN_PREMIUM_PCT > 0 and not tests.get("min_premium_pct"):
        reasons.append(f"premium {per_yld:.2%} of strike < {PUT_MIN_PREMIUM_PCT:.1%}")
    if PUT_MIN_YIELD_OVER_IV > 0 and not tests.get("yield_vs_iv"):
        reasons.append(f"yield {ann_yld:.1%} < {PUT_MIN_YIELD_OVER_IV:.0%} of IV ({iv:.0%})")
    if PUT_MIN_OTM_OVER_IV > 0 and not tests.get("otm_vs_iv"):
        reasons.append(f"OTM {otm:.1%} < {PUT_MIN_OTM_OVER_IV:.0%} of IV ({iv:.0%})")
    if USE_YIELD_OVER_IV and not tests.get("yield_over_iv"):
        reasons.append(f"yield {ann_yld:.1%} below {yiv:.0%} of IV ({iv:.0%})")
    if USE_TBILL_SPREAD and not tests.get("tbill_spread"):
        reasons.append(f"only {risk_prem:.1%} over T-bill")
    if not tests["dte_window"]:
        reasons.append(f"DTE {dte} outside {DTE_MIN}-{DTE_MAX}")
    if not tests["otm_range"]:
        reasons.append(f"OTM {otm:.1%} outside {_om:.0%}-{OTM_MAX:.0%}")
    if not tests["no_earnings"]:
        reasons.append("spans earnings")
    if USE_IVR and not tests.get("iv_rank"):
        reasons.append("IV Rank <50 or missing")
    score = (ann_yld / (iv ** SCORE_IV_EXP) * (delta_pct ** SCORE_POP_EXP) * ((365.0 / dte) ** SCORE_DTE_EXP)
             if (dte and dte > 0 and iv and iv > 0 and ann_yld == ann_yld and delta_pct == delta_pct)
             else float("nan"))
    return {"OTM_%": otm, "Premium": premium, "PeriodYield_%": per_yld,
            "AnnYield_%": ann_yld, "YieldNeeded_%": needed, "Delta_%": delta_pct,
            "IV": iv, "Value": (round(ann_yld / iv * math.sqrt(dte / 365.0), 2)
                                if iv and dte and dte > 0 else float("nan")),
            "Score": (round(score, 2) if score == score else float("nan")),
            "Tbill_%": tbill, "RiskPrem_%": risk_prem,
            "PASS": all(tests.values()), "Reasons": "; ".join(reasons)}


def evaluate_call(row, spot, dte, earnings_in_window, cost_basis, iv_rank=None, delta=None, otm_min=None):
    strike, premium, iv = row["strike"], row["premium"], row["iv"]
    otm     = (strike - spot) / spot
    per_yld = premium / spot if spot else float("nan")
    ann_yld = per_yld * 365.0 / dte if dte else float("nan")
    cd = delta if delta is not None else bs_call_delta(spot, strike, dte / 365.0, iv, RISK_FREE)
    delta_pct = 1.0 - cd
    tbill    = tbill_for(dte)
    risk_prem = ann_yld - tbill
    needed   = YIELD_HURDLE_BASE - otm
    yiv      = YIELD_OVER_IV_SHORT if dte <= DTE_SHORT_CUTOFF else YIELD_OVER_IV_LONG
    req_yield = tiered_yield_needed(otm) if USE_TIERED_YIELD else MIN_ANN_YIELD
    _om = otm_min if otm_min is not None else OTM_MIN_OTHER
    tests = {
        "pop_target":   POP_MIN <= delta_pct <= POP_MAX,
        "min_yield":    ann_yld >= req_yield,
        "min_period_yield": per_yld >= MIN_PERIOD_YIELD,
        "dte_window":   DTE_MIN <= dte <= DTE_MAX,
        "otm_range":    _om <= otm <= OTM_MAX,
        "no_earnings":  not earnings_in_window,
    }
    if CALL_MIN_OTM_OVER_IV > 0:
        tests["otm_vs_iv"] = bool(iv) and otm >= CALL_MIN_OTM_OVER_IV * iv
    if REQUIRE_STRIKE_ABOVE_COST:
        tests["above_cost"] = (cost_basis is None) or (strike >= cost_basis)
    if USE_YIELD_OVER_IV:
        tests["yield_over_iv"] = ann_yld > yiv * iv
    if USE_TBILL_SPREAD:
        tests["tbill_spread"] = risk_prem >= MIN_RISK_PREMIUM
    if USE_IVR:
        tests["iv_rank"] = (iv_rank is not None) and (iv_rank >= IVR_MIN)
    reasons = []
    if not tests["pop_target"]:
        reasons.append(f"POP {delta_pct:.0%} outside {POP_MIN:.0%}-{POP_MAX:.0%}")
    if not tests["min_yield"]:
        reasons.append(f"yield {ann_yld:.1%} < {req_yield:.0%} needed")
    if not tests["min_period_yield"]:
        reasons.append(f"period yield {per_yld:.1%} < {MIN_PERIOD_YIELD:.0%}")
    if CALL_MIN_OTM_OVER_IV > 0 and not tests.get("otm_vs_iv"):
        reasons.append(f"OTM {otm:.1%} < {CALL_MIN_OTM_OVER_IV:.0%} of IV ({iv:.0%})")
    if USE_YIELD_OVER_IV and not tests.get("yield_over_iv"):
        reasons.append(f"yield {ann_yld:.1%} below {yiv:.0%} of IV ({iv:.0%})")
    if USE_TBILL_SPREAD and not tests.get("tbill_spread"):
        reasons.append(f"only {risk_prem:.1%} over T-bill")
    if not tests["dte_window"]:
        reasons.append(f"DTE {dte} outside {DTE_MIN}-{DTE_MAX}")
    if not tests["otm_range"]:
        reasons.append(f"OTM {otm:.1%} outside {_om:.0%}-{OTM_MAX:.0%}")
    if not tests["no_earnings"]:
        reasons.append("spans earnings")
    if REQUIRE_STRIKE_ABOVE_COST and not tests.get("above_cost"):
        reasons.append(f"strike below cost {cost_basis}")
    if USE_IVR and not tests.get("iv_rank"):
        reasons.append("IV Rank <50 or missing")
    score = (ann_yld / (iv ** SCORE_IV_EXP) * (delta_pct ** SCORE_POP_EXP) * ((365.0 / dte) ** SCORE_DTE_EXP)
             if (dte and dte > 0 and iv and iv > 0 and ann_yld == ann_yld and delta_pct == delta_pct)
             else float("nan"))
    return {"OTM_%": otm, "Premium": premium, "PeriodYield_%": per_yld,
            "AnnYield_%": ann_yld, "YieldNeeded_%": needed, "Delta_%": delta_pct,
            "IV": iv, "Value": (round(ann_yld / iv * math.sqrt(dte / 365.0), 2)
                                if iv and dte and dte > 0 else float("nan")),
            "Score": (round(score, 2) if score == score else float("nan")),
            "Tbill_%": tbill, "RiskPrem_%": risk_prem,
            "PASS": all(tests.values()), "Reasons": "; ".join(reasons)}


def _stockanalysis_earnings_date(symbol):
    """Fallback earnings-date source for when Yahoo comes up empty --
    stockanalysis.com exposes a plain, auth-free JSON endpoint (undocumented,
    but stable) with the same kind of info Yahoo's calendar tries to give:
    a list of past/future earnings entries, confirmed or estimated. Same
    "soonest future date" selection as the Yahoo path above."""
    import requests
    try:
        r = requests.get(f"https://stockanalysis.com/api/symbol/s/{symbol}/earnings",
                         timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        entries = (r.json() or {}).get("data") or []
    except Exception:
        return None
    today = dt.date.today()
    fut = []
    for e in entries:
        try:
            d = dt.date.fromisoformat(e["date"])
        except Exception:
            continue
        if d >= today:
            fut.append(d)
    return min(fut) if fut else None


def get_earnings_date(symbol):
    """Manual EARNINGS_DATES override wins; otherwise best-effort Yahoo (two
    methods + retry), falling back to stockanalysis.com if Yahoo comes up
    empty -- Yahoo/yfinance can be flaky (rate limits, bot detection,
    library-side parsing bugs), so a single source shouldn't be the only
    thing standing between a real catalyst and this screener seeing it.

    NO_EARNINGS_TICKERS (ETFs) short-circuit to None right after the manual
    check -- an ETF structurally never has an earnings date, so the full
    Yahoo (2 attempts, 2 methods each, 0.6s sleep between) + stockanalysis.com
    fallback chain below was previously run anyway, guaranteed to fail every
    time, costing ~15-25s per ETF per cache-miss scan for a result that's
    always None."""
    manual = EARNINGS_DATES.get(symbol)
    if manual:
        try:
            d = dt.date.fromisoformat(manual)
            if d >= dt.date.today():
                return d
        except Exception:
            pass
    if symbol in NO_EARNINGS_TICKERS:
        return None
    tkr = _ticker(symbol)
    today = dt.date.today()
    for attempt in range(2):
        try:
            cal = tkr.get_earnings_dates(limit=8)
            if cal is not None and len(cal):
                fut = [d.date() for d in cal.index.to_pydatetime() if d.date() >= today]
                if fut:
                    return min(fut)
        except Exception:
            pass
        try:
            c = tkr.calendar
            ed = c.get("Earnings Date") if isinstance(c, dict) else None
            if ed:
                ed = ed[0] if isinstance(ed, (list, tuple)) else ed
                if hasattr(ed, "date"):
                    ed = ed.date()
                if isinstance(ed, dt.date) and ed >= today:
                    return ed
        except Exception:
            pass
        _time.sleep(0.6)
    return _stockanalysis_earnings_date(symbol)


def _expirations_in_window(symbol, today):
    out = []
    for exp in td_expirations(symbol):
        try:
            d = dt.date.fromisoformat(exp)
        except Exception:
            continue
        dte = (d - today).days
        if DTE_MIN <= dte <= DTE_MAX:
            out.append((exp, d, dte))
    return out


def _best_per_expiration(rows):
    """Keep only the highest-Score contract per (Ticker, Expiration) -- the best
    yield/POP/length balance for that date. Collapses the redundant strike ladder."""
    best = {}
    for r in rows:
        k = (r["Ticker"], r["Expiration"])
        v = r.get("Score")
        v = v if isinstance(v, (int, float)) and v == v else float("-inf")
        b = best.get(k)
        if b is None or v > b[0]:
            best[k] = (v, r)
    return [b[1] for b in best.values()]


def screen_puts(symbol):
    if "put" in open_position_sides(symbol):   # already holding put-side exposure here
        return [], []
    price = td_quote(symbol)
    if not price:
        raise RuntimeError("no quote")
    price = float(price)
    earnings = get_earnings_date(symbol)
    today = dt.date.today()
    passers, near = [], []
    for exp, exp_date, dte in _expirations_in_window(symbol, today):
        earn_win = earnings_blocks(symbol, earnings, today, exp_date)
        for o in td_chain(symbol, exp):
            if o["type"] != "put" or o["strike"] >= price:
                continue
            bid = o["bid"] or 0
            if bid <= 0:
                continue
            if MIN_OPEN_INTEREST > 0 and (o.get("oi") or 0) < MIN_OPEN_INTEREST:
                continue
            premium = bid if PREMIUM_BASIS == "bid" else (bid + (o["ask"] or 0)) / 2
            res = evaluate_put({"strike": o["strike"], "premium": float(premium),
                                "iv": float(o["iv"] or 0)}, price, dte, earn_win,
                               delta=o["delta"], otm_min=otm_min_for(symbol),
                               is_index=(symbol in INDEX_TICKERS))
            _apr = avg_premium_range(symbol, (price - o["strike"]) / price, dte, iv=float(o["iv"] or 0))
            rec = {"Ticker": symbol, "CurrentPrice": round(price, 2), "Strike": o["strike"],
                   "Expiration": exp, "DTE": dte, "EarningsDate": earnings,
                   "AvgPremium": (f"{_apr[0]*100:.1f}%-{_apr[1]*100:.1f}%" if _apr else "-"),
                   "MaxLoss": round(o["strike"] - premium, 2),
                   "# of contracts": contracts_for_target(o["strike"] * 100),
                   **_liq(o), **res}
            if res["PASS"]:
                passers.append(rec)
            elif res["Reasons"].count(";") == 0:
                near.append(rec)
    if PUT_BEST_PER_EXPIRATION:
        passers = _best_per_expiration(passers)
    return passers, near


def screen_calls(symbol, cost_basis):
    if "call" in open_position_sides(symbol):   # already holding call-side exposure here
        return [], []
    price = td_quote(symbol)
    if not price:
        raise RuntimeError("no quote")
    price = float(price)
    earnings = get_earnings_date(symbol)
    today = dt.date.today()
    passers, near = [], []
    for exp, exp_date, dte in _expirations_in_window(symbol, today):
        earn_win = earnings_blocks(symbol, earnings, today, exp_date)
        for o in td_chain(symbol, exp):
            if o["type"] != "call" or o["strike"] <= price:
                continue
            bid = o["bid"] or 0
            if bid <= 0:
                continue
            if MIN_OPEN_INTEREST > 0 and (o.get("oi") or 0) < MIN_OPEN_INTEREST:
                continue
            premium = bid if PREMIUM_BASIS == "bid" else (bid + (o["ask"] or 0)) / 2
            res = evaluate_call({"strike": o["strike"], "premium": float(premium),
                                 "iv": float(o["iv"] or 0)}, price, dte, earn_win,
                                cost_basis, delta=o["delta"], otm_min=otm_min_for(symbol))
            _apr = avg_premium_range(symbol, (o["strike"] - price) / price, dte, "call", iv=float(o["iv"] or 0))
            rec = {"Ticker": symbol, "CurrentPrice": round(price, 2), "CostBasis": cost_basis,
                   "Strike": o["strike"], "Expiration": exp, "DTE": dte,
                   "EarningsDate": earnings,
                   "AvgPremium": (f"{_apr[0]*100:.1f}%-{_apr[1]*100:.1f}%" if _apr else "-"),
                   "MaxLoss": (round(cost_basis - premium, 2) if cost_basis is not None else float("nan")),
                   "# of contracts": contracts_for_target(price * 100),
                   **_liq(o), **res}
            if res["PASS"]:
                passers.append(rec)
            elif res["Reasons"].count(";") == 0:
                near.append(rec)
    if CALL_BEST_PER_EXPIRATION:
        passers = _best_per_expiration(passers)
    return passers, near


def lookup_contracts(symbol, kind="put", strike_min=None, strike_max=None,
                     exp_start=None, exp_end=None):
    """Manual lookup: return EVERY contract for a ticker within the given strike
    and expiration range, with all the same stats, regardless of whether it passes
    the screener's criteria. No PASS filter, no OI floor, no earnings/dedup."""
    price = td_quote(symbol)
    if not price:
        raise RuntimeError("no quote")
    price = float(price)
    earnings = get_earnings_date(symbol)
    today = dt.date.today()
    idx = symbol in INDEX_TICKERS
    rows = []
    for exp in td_expirations(symbol):
        try:
            d = dt.date.fromisoformat(exp)
        except Exception:
            continue
        dte = (d - today).days
        if dte < 0:
            continue
        if exp_start and d < exp_start:
            continue
        if exp_end and d > exp_end:
            continue
        earn_win = earnings_blocks(symbol, earnings, today, d)
        for o in td_chain(symbol, exp):
            if o["type"] != kind:
                continue
            k = o["strike"]
            # sell side only: OTM puts (below price) for CSPs, OTM calls (above price) for covered calls
            if kind == "put" and k >= price:
                continue
            if kind == "call" and k <= price:
                continue
            if strike_min is not None and k < strike_min:
                continue
            if strike_max is not None and k > strike_max:
                continue
            bid = o["bid"] or 0
            if bid <= 0:
                continue
            premium = bid if PREMIUM_BASIS == "bid" else (bid + (o["ask"] or 0)) / 2
            if kind == "put":
                res = evaluate_put({"strike": k, "premium": float(premium),
                                    "iv": float(o["iv"] or 0)}, price, dte, earn_win,
                                   delta=o["delta"], otm_min=otm_min_for(symbol), is_index=idx)
                _apr = avg_premium_range(symbol, (price - k) / price, dte, "put", iv=float(o["iv"] or 0))
                rec = {"Ticker": symbol, "CurrentPrice": round(price, 2), "Strike": k,
                       "Expiration": exp, "DTE": dte, "EarningsDate": earnings,
                       "AvgPremium": (f"{_apr[0]*100:.1f}%-{_apr[1]*100:.1f}%" if _apr else "-"),
                       "MaxLoss": round(k - premium, 2),
                       "# of contracts": contracts_for_target(k * 100),
                       **_liq(o), **res}
            else:
                res = evaluate_call({"strike": k, "premium": float(premium),
                                     "iv": float(o["iv"] or 0)}, price, dte, earn_win,
                                    None, delta=o["delta"], otm_min=otm_min_for(symbol))
                _apr = avg_premium_range(symbol, (k - price) / price, dte, "call", iv=float(o["iv"] or 0))
                rec = {"Ticker": symbol, "CurrentPrice": round(price, 2), "Strike": k,
                       "Expiration": exp, "DTE": dte, "EarningsDate": earnings,
                       "AvgPremium": (f"{_apr[0]*100:.1f}%-{_apr[1]*100:.1f}%" if _apr else "-"),
                       "MaxLoss": float("nan"),  # no cost basis in Contract Lookup -- can't bound the loss
                       "# of contracts": contracts_for_target(price * 100),
                       **_liq(o), **res}
            rows.append(rec)
    return rows


# Lookup shows the same columns as puts (no CostBasis, since it's not tied to a holding)
LOOKUP_COLS = ["Ticker", "CurrentPrice", "Strike", "Expiration", "DTE", "OTM_%", "Premium", "Ask",
               "AvgPremium", "AnnYield_%", "PeriodYield_%", "Delta_%", "IV", "Score",
               "# of contracts", "MaxLoss", "OpenInt", "Volume", "EarningsDate"]


PUT_COLS = ["Ticker", "CurrentPrice", "Strike", "Expiration", "DTE", "OTM_%", "Premium", "Ask",
            "AvgPremium", "AnnYield_%", "PeriodYield_%", "Delta_%", "IV", "Score",
            "# of contracts", "MaxLoss", "OpenInt", "Volume", "EarningsDate"]
CALL_COLS = ["Ticker", "CurrentPrice", "CostBasis", "Strike", "Expiration", "DTE", "OTM_%",
             "Premium", "Ask", "AvgPremium", "AnnYield_%", "PeriodYield_%", "Delta_%", "IV", "Score",
             "# of contracts", "MaxLoss", "OpenInt", "Volume", "EarningsDate"]
PCT_COLS = {"OTM_%", "PeriodYield_%", "AnnYield_%", "Delta_%",
            "Tbill_%", "RiskPrem_%", "IV"}


def _df(rows, cols, sort_by=("Ticker", "Strike"), asc=(True, False)):
    if not rows:
        return pd.DataFrame(columns=cols)
    # default: ticker A->Z, then strike high->low. Puts pass ("Ticker","Delta_%")
    # with asc=(True, False) to rank each ticker safest-first (highest POP).
    return pd.DataFrame(rows).sort_values(list(sort_by),
                                          ascending=list(asc))[cols]


def write_report(sheets, settings, path=OUTPUT_FILE):
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    green = PatternFill("solid", fgColor="C6EFCE")
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, df in sheets + [("Settings", settings)]:
            df.to_excel(xw, sheet_name=name, index=False)
            sh = xw.sheets[name]
            for ci, col in enumerate(df.columns, 1):
                c = sh.cell(row=1, column=ci)
                c.fill, c.font = hdr_fill, hdr_font
                c.alignment = Alignment(horizontal="center")
                vals = [len(str(col))] + [len(str(v)) for v in df[col].head(30)]
                sh.column_dimensions[c.column_letter].width = max(10, min(24, max(vals) + 2))
                if col in PCT_COLS:
                    for ri in range(2, len(df) + 2):
                        sh.cell(row=ri, column=ci).number_format = "0.0%"
            sh.freeze_panes = "A2"
            if len(df):
                sh.auto_filter.ref = sh.dimensions
            if name != "Settings":
                for ri in range(2, len(df) + 2):
                    for ci in range(1, len(df.columns) + 1):
                        sh.cell(row=ri, column=ci).fill = green


def _fmt(df):
    d = df.copy()
    for c in PCT_COLS:
        if c in d.columns:
            d[c] = (d[c] * 100).round(1).astype(str) + "%"
    # Premium: "$bid-$ask (total range across the # of contracts that reach the cash target)".
    # Showing the live bid-ask spread here folds in what the old separate Spread_$
    # liquidity column used to convey, so that column is gone.
    if "Premium" in d.columns and "# of contracts" in d.columns:
        def _prem(bid, ask, n):
            if pd.isna(bid):
                return "-"
            ask = ask if pd.notna(ask) else bid
            if pd.isna(n):
                return f"${bid:.2f}-${ask:.2f}"
            return f"${bid:.2f}-${ask:.2f} (${bid * 100 * int(n):,.0f}-${ask * 100 * int(n):,.0f})"
        ask_col = d["Ask"] if "Ask" in d.columns else d["Premium"]
        d["Premium"] = [_prem(p, a, n) for p, a, n in zip(d["Premium"], ask_col, d["# of contracts"])]
        if "Ask" in d.columns:
            d = d.drop(columns=["Ask"])
    elif "Premium" in d.columns:
        d["Premium"] = "$" + d["Premium"].round(2).astype(str)
    for c in ("CostBasis", "CurrentPrice"):
        if c in d.columns:
            d[c] = "$" + d[c].round(2).astype(str)
    # MaxLoss: "$/share (total across the # of contracts that reach the cash target)", same as Premium
    if "MaxLoss" in d.columns and "# of contracts" in d.columns:
        def _mloss(v, n):
            if pd.isna(v):
                return "-"
            if pd.isna(n):
                return f"${v:.2f}"
            return f"${v:.2f} (${v * 100 * int(n):,.0f})"
        d["MaxLoss"] = [_mloss(v, n) for v, n in zip(d["MaxLoss"], d["# of contracts"])]
    elif "MaxLoss" in d.columns:
        d["MaxLoss"] = d["MaxLoss"].apply(lambda v: f"${v:.2f}" if pd.notna(v) else "-")
    return d


def _tbl(df, empty):
    return _fmt(df).to_html(index=False, classes="q", border=0) if len(df) else f"<p>{empty}</p>"


def write_html(put_pass, call_pass, vix=None, path="qualifying_contracts.html"):
    style = ("<style>body{font-family:Arial,Helvetica,sans-serif;margin:24px;"
             "background:#0f1115;color:#e8e8e8}h1{font-size:20px}h2{margin-top:30px;"
             "border-bottom:2px solid #1F3864;padding-bottom:4px}"
             "p{color:#aaa}table{border-collapse:collapse;width:100%;font-size:13px;margin-top:8px}"
             "th,td{border:1px solid #333;padding:6px 9px;text-align:right}"
             "th{background:#1F3864;color:#fff}td:first-child,th:first-child{text-align:left}"
             ".q tbody td{background:#12331d}</style>")
    vtxt = f"VIX {vix} ({vix_regime(vix)})" if vix is not None else "VIX n/a"
    p = ["<html><head><meta charset='utf-8'>", style, "</head><body>",
         "<h1>Wheel Screener</h1>",
         f"<p>{vtxt}. Generated {dt.date.today()}. "
         "Delta_% = chance of keeping the premium (1 - delta). "
         "YieldNeeded_% = 25% - OTM%. Not financial advice.</p>",
         "<h2>Cash-Secured Puts</h2>",
         _tbl(put_pass, "None - nothing pays enough; stay in T-bills."),
         "<h2>Covered Calls (strike above your cost)</h2>",
         _tbl(call_pass, "None - no call above your cost pays enough."),
         "</body></html>"]
    with open(path, "w", encoding="utf-8") as f:
        f.write("".join(p))


def main():
    vix = get_vix()
    pp = []
    for t in PUT_TICKERS:
        try:
            a, _ = screen_puts(t)
            pp += a
            print(f"PUT  {t}: {len(a)} qualifying")
        except Exception as e:
            print(f"PUT  {t}: ERROR {e}")
    cp = []
    for t, cost in HOLDINGS.items():
        try:
            a, _ = screen_calls(t, cost)
            cp += a
            print(f"CALL {t}: {len(a)} qualifying")
        except Exception as e:
            print(f"CALL {t}: ERROR {e}")

    put_pass = _df(pp, PUT_COLS)
    call_pass = _df(cp, CALL_COLS)
    settings = pd.DataFrame({
        "Setting": ["POP min", "POP max", "DTE min", "DTE max",
                    "Yield needed base (25-OTM)", "Use T-bill spread", "Min risk premium",
                    "Use IV Rank", "Premium basis", "Current VIX", "VIX regime", "Run date"],
        "Value": [POP_MIN, POP_MAX, DTE_MIN, DTE_MAX, YIELD_HURDLE_BASE,
                  USE_TBILL_SPREAD, MIN_RISK_PREMIUM, USE_IVR, PREMIUM_BASIS,
                  vix, vix_regime(vix), dt.date.today().isoformat()]})
    write_report([("Puts", put_pass), ("Covered Calls", call_pass)], settings, OUTPUT_FILE)
    write_html(put_pass, call_pass, vix, "qualifying_contracts.html")
    print(f"\nVIX {vix} ({vix_regime(vix)}). Puts: {len(put_pass)} | Calls: {len(call_pass)}.")


if __name__ == "__main__":
    main()
